from openpyxl import load_workbook
import traceback
from datetime import datetime
import os

try:
    # Cargar el nombre del archivo xlsx y de la hoja
    carpeta_archivos = "datos_excel"
    archivo_excel = os.path.join(carpeta_archivos, "REPORTE-DE-PERSONAL.xlsx")
    hoja = "NM3670ATXT"

    # Cargar el libro de trabajo y seleccionar la hoja
    libro_trabajo = load_workbook(archivo_excel)
    hoja_datos = libro_trabajo[hoja]

    # Verificar la cantidad de columnas disponibles
    num_columnas = hoja_datos.max_column
    print(f"La hoja '{hoja}' tiene {num_columnas} columnas.")

    # Obtener los nombres de las columnas por su orden respectivo
    nombres_columnas = [hoja_datos.cell(row=1, column=i).value for i in range(1, num_columnas + 1)]
    print("Nombres de las columnas por su orden respectivo:")
    print(nombres_columnas)

    # Validar la cantidad de columnas
    if num_columnas != 5:
        raise ValueError(f"Error: El número de columnas es {num_columnas}. Deben ser 5 columnas.")

    # Validar el orden y nombres de las columnas
    columnas_esperadas = ["COMPANIA", "CEDULA", "APELLIDOS", "NOMBRES", "CARGO"]
    if nombres_columnas != columnas_esperadas:
        raise ValueError(f"Error: El orden o los nombres de las columnas no son correctos.")

    # Columnas que se desea obtener
    columnas_seleccionadas = [i for i in range(1, num_columnas + 1)]

    # Obtener todas las filas de la hoja
    todas_las_filas = list(hoja_datos.iter_rows(min_row=2, max_row=hoja_datos.max_row, min_col=1, max_col=num_columnas))

    # Escribir los datos seleccionados en un archivo txt
    exportar_datos = os.path.join(carpeta_archivos, 'cargos.txt')
    with open(exportar_datos, 'w') as archivo_txt:
        for fila in todas_las_filas:
            valores_fila = [str(fila[col - 1].value) for col in columnas_seleccionadas]
            linea = '\t'.join(valores_fila)
            archivo_txt.write(linea + '\n')

    print(f"Datos exportados correctamente a '{exportar_datos}'.")
except Exception as e:
    # Capturar cualquier excepción y exportar el mensaje de error
    mensaje_error = f"Error: {str(e)}\n"
    mensaje_error += traceback.format_exc()

    carpeta_errores = "errores_extractor"
    os.makedirs(carpeta_errores, exist_ok=True)

    fecha_error = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    archivo_error = os.path.join(carpeta_errores, f'error_extractor_{fecha_error}.txt')

    with open(archivo_error, 'w') as archivo_err:
        archivo_err.write(mensaje_error)

    print(f"Se ha producido un error. Detalles en '{archivo_error}'.")
